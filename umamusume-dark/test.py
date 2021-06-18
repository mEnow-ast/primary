import os
import glob
import sys
import datetime
import tkinter
import openpyxl as op
import subprocess as sp
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox


def is_empty(cell) :
    """
    cell : 空欄か判断するセル

    """
    return cell.value is None or not str(cell.value).strip()


def cast_cereal(date):
    """
    date : シリアル値へ変更したい日付(str)

    """
    orderday = date[0:8]
    dt = datetime.datetime.strptime(orderday, '%Y%m%d')- datetime.datetime(1899, 12, 31)
    return dt.days + 1


def merge_month_exl(folder_path, file_path):
    """
    folder_path : 処理をしたいフォルダ
    file_path : 排出されるファイル名

    Excelの日付と合致しないときは、下記どちらかの処理をしてください。
    １）Excel：「ファイル」→「オプション」→「詳細設定」→「1904年から計算する」のチェックを外す
    ２）「serial = dt.days + 1」を「serial = dt.days」にする

    """
    # テンプレート読み込み
    lb = op.load_workbook(r"temp.xlsx")
    print(lb)
    ls = lb.worksheets[0]
    leng = 4 #4行名からスタートする 件数カウントは1からなので-3している

    # フォルダ内のファイル一覧を取得
    exl_dir = glob.glob(folder_path+"\起算日確認_*")

    # 変換用リプレイス文字
    paht_head = folder_path + r"\起算日確認_"

    # 集計シートへ書き込み
    for i,sheet in enumerate(exl_dir) :
        try :
            wb = op.load_workbook(sheet)
        except Exception as a:
            print(a)
            return "ZIP"
        ws = wb.worksheets[0]
        strs = exl_dir[i].replace(paht_head, "") #パス等を消し、日付を先頭にする

        if "翌日" in strs : # 翌日Excelの処理
            serial = cast_cereal(strs) +1
            for y,rows in enumerate(ws.iter_rows(min_row=2)) : # 書き込み処理
                if all(is_empty(c) for c in rows): # 行すべてが空白であればfor文を終える
                    break
                ls.cell(row=leng, column=1).value = leng - 3 # 件数カウント
                ls.cell(row=leng, column=2).value = ws.cell(row=y+2, column=1).value # 基地局ID
                ls.cell(row=leng, column=3).value = ws.cell(row=y+2, column=2).value # 基地局名称
                ls.cell(row=leng, column=4).value = serial # 日付
                ls.cell(row=leng, column=5).value = "翌日チェック" # 作業日
                leng += 1 # データの長さを+1
        else : # 当日Excelの処理
            serial = cast_cereal(strs)
            for y,rows in enumerate(ws.iter_rows(min_row=2)) : # 書き込み処理
                if all(is_empty(c) for c in rows): # 行すべてが空白であればfor文を終える
                    break
                ls.cell(row=leng, column=1).value = leng - 3 #件数カウント
                ls.cell(row=leng, column=2).value = ws.cell(row=y+2, column=1).value # 基地局ID
                ls.cell(row=leng, column=3).value = ws.cell(row=y+2, column=2).value # 基地局名称
                ls.cell(row=leng, column=4).value = serial #日付
                ls.cell(row=leng, column=5).value = "当日チェック" # 作業日
                leng += 1 # データの長さを+1

    # 注文書番号の記載
    orderday = exl_dir[-1].replace(paht_head, "")
    if "2021" in orderday :
        month = int(orderday[4:6])
        order_id = 6 + month * 2
        orders = "注文書番号 ：M000538228-"  + str(order_id-1) + "、M000538228-" + str(order_id)
        ls.cell(row=1,column=4).value = orders
    else :
        month = int(orderday[4:6])
        order_id = 30 + month * 2
        orders = "注文書番号 ：M000538228-"  + str(order_id-1) + "、M000538228-" + str(order_id)
        ls.cell(row=1,column=4).value = orders

    # 年月日書き込み
    orderday = orderday[0:8]
    dt_now = datetime.datetime.strptime(orderday, '%Y%m%d')
    today = dt_now.strftime('%Y年 %m月') + "分 三技協・作業報告書（別紙5）"
    ls.cell(row=1,column=2).value = today

    # ファイルの保存
    lb.save(file_path)

    #結果フラグ
    return "complete"


def ask_folder():
    """
    参照ボタンの動作

    """
    path = filedialog.askdirectory()
    folder_path.set(path)
    print(path)


def ask_file():
    """
    選択ボタンの動作

    """
    path = filedialog.asksaveasfilename(filetypes=[("excel", "*.xlsx")], defaultextension=".xlsx")

    file_path.set(path)
    print(path)



def close_window():
    """
    主ポップアップを閉じる

    """
    main_win.destroy()


def app():
    """
    実行ボタンの動作

    """
    input_dir = folder_path.get()
    # 保存するexcelファイルを指定
    file_name = file_path.get()
    # 両方に値がない場合は実行しない
    if not input_dir or not file_name:
        return
    # [同じ階層に作成する]がチェックされている場合、前パスを付ける
    if same_check.get() == True :
        if input_dir not in file_name :
            output_file = input_dir + r"/" + file_name
        else :
            output_file = file_name
    else :
        if r":" in file_name :
            output_file = file_name
        else :
            check = messagebox.showinfo("エラー", "予期せぬエラーが発生しました。\nパスが正しいかどうか確認してください。")
            if check == "ok" : return
    # 拡張子がない場合、付ける
    if ".xlsx" not in output_file :
        output_file += ".xlsx"

    # 結合実行 ファイル作成の可否でポップアップを表示
    result = merge_month_exl(input_dir, output_file)
    if result == "complete" :
        check = messagebox.askquestion("完了", "完了しました。\n作成したフォルダを表示しますか？")
        if check == "yes" :
            dirname = os.path.dirname(output_file) #フォルダ名を取得
            sp.Popen(['explorer', dirname.replace("/", "\\")], shell=True)
            close_window()
        else : close_window()
    elif result == "ZIP" :
        check = messagebox.showinfo("エラー", "予期せぬエラーが発生しました。\nファイルの権限が撤廃されているか確認してください。")
        if check == "ok" : return



if __name__ == "__main__" :
    """
    主ウィンドウの作成・表示

    """

    # メインウィンドウ
    main_win = tkinter.Tk()
    main_win.title("月末集計")
    main_win.geometry("500x150")

    # メインフレーム
    main_frm = ttk.Frame(main_win)
    main_frm.grid(column=0, row=0, sticky=tkinter.NSEW, padx=5, pady=10)

    #パラメータ
    folder_path = tkinter.StringVar()
    file_path = tkinter.StringVar()
    same_check = tkinter.BooleanVar()

    # ウィジェット（フォルダ名）
    folder_label = ttk.Label(main_frm, text="フォルダ指定")
    folder_box = ttk.Entry(main_frm, textvariable=folder_path)
    folder_btn = ttk.Button(main_frm, text="参照", command=ask_folder)

    # ウィジェット（保存先）
    file_label = ttk.Label(main_frm, text="ファイル名")
    file_box = ttk.Entry(main_frm, textvariable=file_path)
    file_btn = ttk.Button(main_frm, text="選択", command=ask_file)

    # ウィジェット作成（階層チェックボックス）
    same_folder = ttk.Checkbutton(main_frm, var=same_check,
                                  text="選択したフォルダと同じ階層に出力する（ファイル名のみ入力する）")

    # ウィジェット（実行ボタン）
    app_btn = ttk.Button(main_frm, text="実行", command=app)

    # ウィジェットの配置
    folder_label.grid(column=0, row=0, pady=10)
    folder_box.grid(column=1, row=0, sticky=tkinter.EW, padx=5)
    folder_btn.grid(column=2, row=0)
    file_label.grid(column=0, row=1, pady=10)
    file_box.grid(column=1, row=1, sticky=tkinter.EW, padx=5)
    file_btn.grid(column=2, row=1)
    same_folder.grid(column=1, row=2)
    app_btn.grid(column=1, row=3)

    # 配置設定
    main_win.columnconfigure(0, weight=1)
    main_win.rowconfigure(0, weight=1)
    main_frm.columnconfigure(1, weight=1)

    # 描画開始
    main_win.mainloop()import os
import glob
import sys
import datetime
import tkinter
import openpyxl as op
import subprocess as sp
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox


def is_empty(cell) :
    """
    cell : 空欄か判断するセル

    """
    return cell.value is None or not str(cell.value).strip()


def cast_cereal(date):
    """
    date : シリアル値へ変更したい日付(str)

    """
    orderday = date[0:8]
    dt = datetime.datetime.strptime(orderday, '%Y%m%d')- datetime.datetime(1899, 12, 31)
    return dt.days + 1


def merge_month_exl(folder_path, file_path):
    """
    folder_path : 処理をしたいフォルダ
    file_path : 排出されるファイル名

    Excelの日付と合致しないときは、下記どちらかの処理をしてください。
    １）Excel：「ファイル」→「オプション」→「詳細設定」→「1904年から計算する」のチェックを外す
    ２）「serial = dt.days + 1」を「serial = dt.days」にする

    """
    # テンプレート読み込み
    lb = op.load_workbook(r"temp.xlsx")
    print(lb)
    ls = lb.worksheets[0]
    leng = 4 #4行名からスタートする 件数カウントは1からなので-3している

    # フォルダ内のファイル一覧を取得
    exl_dir = glob.glob(folder_path+"\起算日確認_*")

    # 変換用リプレイス文字
    paht_head = folder_path + r"\起算日確認_"

    # 集計シートへ書き込み
    for i,sheet in enumerate(exl_dir) :
        try :
            wb = op.load_workbook(sheet)
        except Exception as a:
            print(a)
            return "ZIP"
        ws = wb.worksheets[0]
        strs = exl_dir[i].replace(paht_head, "") #パス等を消し、日付を先頭にする

        if "翌日" in strs : # 翌日Excelの処理
            serial = cast_cereal(strs) +1
            for y,rows in enumerate(ws.iter_rows(min_row=2)) : # 書き込み処理
                if all(is_empty(c) for c in rows): # 行すべてが空白であればfor文を終える
                    break
                ls.cell(row=leng, column=1).value = leng - 3 # 件数カウント
                ls.cell(row=leng, column=2).value = ws.cell(row=y+2, column=1).value # 基地局ID
                ls.cell(row=leng, column=3).value = ws.cell(row=y+2, column=2).value # 基地局名称
                ls.cell(row=leng, column=4).value = serial # 日付
                ls.cell(row=leng, column=5).value = "翌日チェック" # 作業日
                leng += 1 # データの長さを+1
        else : # 当日Excelの処理
            serial = cast_cereal(strs)
            for y,rows in enumerate(ws.iter_rows(min_row=2)) : # 書き込み処理
                if all(is_empty(c) for c in rows): # 行すべてが空白であればfor文を終える
                    break
                ls.cell(row=leng, column=1).value = leng - 3 #件数カウント
                ls.cell(row=leng, column=2).value = ws.cell(row=y+2, column=1).value # 基地局ID
                ls.cell(row=leng, column=3).value = ws.cell(row=y+2, column=2).value # 基地局名称
                ls.cell(row=leng, column=4).value = serial #日付
                ls.cell(row=leng, column=5).value = "当日チェック" # 作業日
                leng += 1 # データの長さを+1

    # 注文書番号の記載
    orderday = exl_dir[-1].replace(paht_head, "")
    if "2021" in orderday :
        month = int(orderday[4:6])
        order_id = 6 + month * 2
        orders = "注文書番号 ：M000538228-"  + str(order_id-1) + "、M000538228-" + str(order_id)
        ls.cell(row=1,column=4).value = orders
    else :
        month = int(orderday[4:6])
        order_id = 30 + month * 2
        orders = "注文書番号 ：M000538228-"  + str(order_id-1) + "、M000538228-" + str(order_id)
        ls.cell(row=1,column=4).value = orders

    # 年月日書き込み
    orderday = orderday[0:8]
    dt_now = datetime.datetime.strptime(orderday, '%Y%m%d')
    today = dt_now.strftime('%Y年 %m月') + "分 三技協・作業報告書（別紙5）"
    ls.cell(row=1,column=2).value = today

    # ファイルの保存
    lb.save(file_path)

    #結果フラグ
    return "complete"


def ask_folder():
    """
    参照ボタンの動作

    """
    path = filedialog.askdirectory()
    folder_path.set(path)
    print(path)


def ask_file():
    """
    選択ボタンの動作

    """
    path = filedialog.asksaveasfilename(filetypes=[("excel", "*.xlsx")], defaultextension=".xlsx")

    file_path.set(path)
    print(path)



def close_window():
    """
    主ポップアップを閉じる

    """
    main_win.destroy()


def app():
    """
    実行ボタンの動作

    """
    input_dir = folder_path.get()
    # 保存するexcelファイルを指定
    file_name = file_path.get()
    # 両方に値がない場合は実行しない
    if not input_dir or not file_name:
        return
    # [同じ階層に作成する]がチェックされている場合、前パスを付ける
    if same_check.get() == True :
        if input_dir not in file_name :
            output_file = input_dir + r"/" + file_name
        else :
            output_file = file_name
    else :
        if r":" in file_name :
            output_file = file_name
        else :
            check = messagebox.showinfo("エラー", "予期せぬエラーが発生しました。\nパスが正しいかどうか確認してください。")
            if check == "ok" : return
    # 拡張子がない場合、付ける
    if ".xlsx" not in output_file :
        output_file += ".xlsx"

    # 結合実行 ファイル作成の可否でポップアップを表示
    result = merge_month_exl(input_dir, output_file)
    if result == "complete" :
        check = messagebox.askquestion("完了", "完了しました。\n作成したフォルダを表示しますか？")
        if check == "yes" :
            dirname = os.path.dirname(output_file) #フォルダ名を取得
            sp.Popen(['explorer', dirname.replace("/", "\\")], shell=True)
            close_window()
        else : close_window()
    elif result == "ZIP" :
        check = messagebox.showinfo("エラー", "予期せぬエラーが発生しました。\nファイルの権限が撤廃されているか確認してください。")
        if check == "ok" : return



if __name__ == "__main__" :
    """
    主ウィンドウの作成・表示

    """

    # メインウィンドウ
    main_win = tkinter.Tk()
    main_win.title("月末集計")
    main_win.geometry("500x150")

    # メインフレーム
    main_frm = ttk.Frame(main_win)
    main_frm.grid(column=0, row=0, sticky=tkinter.NSEW, padx=5, pady=10)

    #パラメータ
    folder_path = tkinter.StringVar()
    file_path = tkinter.StringVar()
    same_check = tkinter.BooleanVar()

    # ウィジェット（フォルダ名）
    folder_label = ttk.Label(main_frm, text="フォルダ指定")
    folder_box = ttk.Entry(main_frm, textvariable=folder_path)
    folder_btn = ttk.Button(main_frm, text="参照", command=ask_folder)

    # ウィジェット（保存先）
    file_label = ttk.Label(main_frm, text="ファイル名")
    file_box = ttk.Entry(main_frm, textvariable=file_path)
    file_btn = ttk.Button(main_frm, text="選択", command=ask_file)

    # ウィジェット作成（階層チェックボックス）
    same_folder = ttk.Checkbutton(main_frm, var=same_check,
                                  text="選択したフォルダと同じ階層に出力する（ファイル名のみ入力する）")

    # ウィジェット（実行ボタン）
    app_btn = ttk.Button(main_frm, text="実行", command=app)

    # ウィジェットの配置
    folder_label.grid(column=0, row=0, pady=10)
    folder_box.grid(column=1, row=0, sticky=tkinter.EW, padx=5)
    folder_btn.grid(column=2, row=0)
    file_label.grid(column=0, row=1, pady=10)
    file_box.grid(column=1, row=1, sticky=tkinter.EW, padx=5)
    file_btn.grid(column=2, row=1)
    same_folder.grid(column=1, row=2)
    app_btn.grid(column=1, row=3)

    # 配置設定
    main_win.columnconfigure(0, weight=1)
    main_win.rowconfigure(0, weight=1)
    main_frm.columnconfigure(1, weight=1)

    # 描画開始
    main_win.mainloop()
